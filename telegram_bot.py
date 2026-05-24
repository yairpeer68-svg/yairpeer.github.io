"""
╔══════════════════════════════════════════════════════════════════╗
║       TELEGRAM ULTRA SEARCH BOT — v3.0                         ║
║  Relevance · Health Score · Compare · Alerts · Categories      ║
║  PDF Export · Dashboard · Top Users · Share · Proxy Pool       ║
╚══════════════════════════════════════════════════════════════════╝
"""

from telethon import TelegramClient, events, functions, types, Button
from telethon.tl.functions.channels import GetFullChannelRequest
from telethon.tl.functions.contacts import ResolveUsernameRequest
from telethon.tl.functions.messages import GetHistoryRequest
from telethon.errors import FloodWaitError
import asyncio
import aiohttp
import aiosqlite
import re
import urllib.parse
import json
import logging
import os
import csv
import io
import traceback
import platform
import psutil
from datetime import datetime, timedelta
from dotenv import load_dotenv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

try:
    import redis.asyncio as aioredis
    HAS_REDIS = True
except ImportError:
    HAS_REDIS = False

load_dotenv()

# ─────────────────────────────────────────────
#  לוגים
# ─────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('bot.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────
#  ⬇️  שנה כאן את הפרטים שלך
# ─────────────────────────────────────────────
API_ID    = 31798000
API_HASH  = '8b930d39b76e44e98b45237c8d63db0f'
BOT_TOKEN = '8616693129:AAFiPTc3pKQrUKhsolRxGhNJW1TY85_Y3Kw'
ADMIN_IDS: set[int] = {7670817500}
# ─────────────────────────────────────────────
ALLOWED_USERS: set[int] = set()

RATE_LIMIT_PER_HOUR = 20
CACHE_TTL_SECONDS   = 3600
SEARCH_VARIANTS     = 15
BOT_START_TIME      = datetime.utcnow()
ENRICH_SEM: asyncio.Semaphore | None = None  # initialized in run_all()

# Proxy pool (optional) — PROXIES=http://user:pass@host:port,http://...
PROXY_LIST: list[str] = [p.strip() for p in os.getenv('PROXIES', '').split(',') if p.strip()]
_proxy_idx = 0

def _next_proxy() -> str | None:
    global _proxy_idx
    if not PROXY_LIST:
        return None
    proxy = PROXY_LIST[_proxy_idx % len(PROXY_LIST)]
    _proxy_idx += 1
    return proxy

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/124.0.0.0 Safari/537.36'
    )
}

# Redis (optional) — REDIS_URL=redis://localhost:6379
_redis_client = None

async def _get_redis():
    global _redis_client
    if HAS_REDIS and _redis_client is None:
        try:
            _redis_client = aioredis.from_url(os.getenv('REDIS_URL', 'redis://localhost:6379'))
            await _redis_client.ping()
            log.info("Redis connected ✅")
        except Exception as e:
            log.warning(f"Redis unavailable, falling back to SQLite: {e}")
            _redis_client = None
    return _redis_client

# ─────────────────────────────────────────────
#  קטגוריות, מילים נרדפות, Blacklist
# ─────────────────────────────────────────────
CATEGORIES: dict[str, list[str]] = {
    'קריפטו':    ['bitcoin', 'btc', 'crypto', 'ethereum', 'defi', 'nft', 'ביטקוין', 'קריפטו', 'blockchain', 'web3'],
    'מניות':     ['stocks', 'trading', 'forex', 'השקעות', 'בורסה', 'מניות', 'finance', 'investment'],
    'חדשות':     ['news', 'חדשות', 'עדכונים', 'breaking', 'אקטואליה', 'daily'],
    'ספורט':     ['sport', 'football', 'basketball', 'כדורגל', 'ספורט', 'soccer', 'tennis', 'nba'],
    'טכנולוגיה': ['tech', 'technology', 'ai', 'programming', 'code', 'טכנולוגיה', 'dev', 'software', 'cyber'],
    'פוליטיקה':  ['politics', 'פוליטיקה', 'ממשלה', 'כנסת', 'government', 'election', 'בחירות'],
    'בידור':     ['entertainment', 'music', 'movies', 'בידור', 'מוזיקה', 'סרטים', 'humor', 'memes'],
    'חינוך':     ['education', 'learn', 'course', 'חינוך', 'למידה', 'קורס', 'study', 'university'],
    'בריאות':    ['health', 'fitness', 'בריאות', 'כושר', 'תזונה', 'diet', 'workout', 'medical'],
    'נדלן':      ['real estate', 'נדלן', 'דירות', 'property', 'apartment', 'housing'],
}

SYNONYMS: dict[str, list[str]] = {
    'ביטקוין':   ['bitcoin', 'btc', 'קריפטו', 'crypto', 'blockchain'],
    'קריפטו':    ['crypto', 'bitcoin', 'ethereum', 'defi', 'nft'],
    'מניות':     ['stocks', 'trading', 'השקעות', 'בורסה', 'forex'],
    'חדשות':     ['news', 'עדכונים', 'breaking'],
    'ספורט':     ['sport', 'כדורגל', 'football', 'basketball'],
    'טכנולוגיה': ['tech', 'technology', 'ai', 'בינה מלאכותית'],
    'פוליטיקה':  ['politics', 'ממשלה', 'כנסת', 'בחירות'],
}

SPAM_KEYWORDS = [
    'free money', 'earn $$$', 'get rich', 'porn', 'xxx', 'nude',
    'הכסף הגדול', 'הרוויח מהיר', 'הימורים', 'קזינו',
]

LANGUAGE_MAP = {
    'he': '🇮🇱', 'en': '🇺🇸', 'ru': '🇷🇺',
    'ar': '🇸🇦', 'de': '🇩🇪', 'fr': '🇫🇷',
    'es': '🇪🇸', 'tr': '🇹🇷', 'uk': '🇺🇦',
}

# ─────────────────────────────────────────────
#  מקורות חיפוש חיצוניים — 30 מקורות
# ─────────────────────────────────────────────
SEARCH_SOURCES: dict[str, callable] = {
    'Google':        lambda q: f"https://www.google.com/search?q={urllib.parse.quote('site:t.me ' + q)}&num=100&hl=he",
    'Bing':          lambda q: f"https://www.bing.com/search?q={urllib.parse.quote('site:t.me ' + q)}&count=50",
    'DDG':           lambda q: f"https://html.duckduckgo.com/html/?q={urllib.parse.quote('site:t.me ' + q)}",
    'Yandex':        lambda q: f"https://yandex.com/search/?text={urllib.parse.quote('site:t.me ' + q)}&lr=10174",
    'Yahoo':         lambda q: f"https://search.yahoo.com/search?p={urllib.parse.quote('site:t.me ' + q)}&n=50",
    'Brave':         lambda q: f"https://search.brave.com/search?q={urllib.parse.quote('site:t.me ' + q)}",
    'Startpage':     lambda q: f"https://www.startpage.com/search?q={urllib.parse.quote('site:t.me ' + q)}",
    'Ecosia':        lambda q: f"https://www.ecosia.org/search?q={urllib.parse.quote('site:t.me ' + q)}",
    'Mojeek':        lambda q: f"https://www.mojeek.com/search?q={urllib.parse.quote('site:t.me ' + q)}",
    'Baidu':         lambda q: f"https://www.baidu.com/s?wd={urllib.parse.quote('site:t.me ' + q)}",
    'TGStat':        lambda q: f"https://tgstat.ru/en/search?q={urllib.parse.quote(q)}",
    'TGStatCom':     lambda q: f"https://tgstat.com/search?q={urllib.parse.quote(q)}",
    'Telemetr':      lambda q: f"https://telemetr.io/en/channels?channel={urllib.parse.quote(q)}",
    'Combot':        lambda q: f"https://combot.org/chats?q={urllib.parse.quote(q)}&lng=&type=",
    'Lyzem':         lambda q: f"https://lyzem.com/search?q={urllib.parse.quote(q)}&lang=",
    'HotTG':         lambda q: f"https://hottg.com/search?q={urllib.parse.quote(q)}",
    'TgPw':          lambda q: f"https://tg.pw/search/{urllib.parse.quote(q)}",
    'Tchannels':     lambda q: f"https://tchannels.me/search?q={urllib.parse.quote(q)}",
    'TGrам':         lambda q: f"https://tgram.ru/search?q={urllib.parse.quote(q)}",
    'TelegramDB':    lambda q: f"https://telegramdb.org/search?q={urllib.parse.quote(q)}",
    'TgDev':         lambda q: f"https://tgdev.io/search?q={urllib.parse.quote(q)}",
    'TgList':        lambda q: f"https://tglist.net/search?q={urllib.parse.quote(q)}",
    'TeleList':      lambda q: f"https://telelist.com/search/{urllib.parse.quote(q)}",
    'TelegramGroup': lambda q: f"https://telegram-group.com/search?q={urllib.parse.quote(q)}",
    'TGChannels':    lambda q: f"https://telegramchannels.me/search?query={urllib.parse.quote(q)}",
    'TGDir':         lambda q: f"https://tg.directory/search?q={urllib.parse.quote(q)}",
    'Tlgrm':         lambda q: f"https://tlgrm.eu/channels/search/{urllib.parse.quote(q)}",
    'SocialBlade':   lambda q: f"https://socialblade.com/telegram/search?q={urllib.parse.quote(q)}",
    'Reddit':        lambda q: f"https://www.reddit.com/search/?q={urllib.parse.quote('telegram t.me ' + q)}&type=link&sort=relevance",
    'GitHub':        lambda q: f"https://github.com/search?q={urllib.parse.quote('t.me ' + q)}&type=code",
}

# ─────────────────────────────────────────────
#  DB — aiosqlite (async)
# ─────────────────────────────────────────────
DB_PATH = 'bot_data.db'

async def init_db():
    async with aiosqlite.connect(DB_PATH) as con:
        await con.executescript("""
            CREATE TABLE IF NOT EXISTS searches (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id     INTEGER NOT NULL,
                query       TEXT    NOT NULL,
                results_cnt INTEGER DEFAULT 0,
                ts          TEXT    DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS cache (
                query      TEXT PRIMARY KEY,
                data       TEXT NOT NULL,
                expires_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS subscriptions (
                user_id    INTEGER NOT NULL,
                query      TEXT    NOT NULL,
                last_count INTEGER DEFAULT 0,
                PRIMARY KEY (user_id, query)
            );
            CREATE TABLE IF NOT EXISTS rate_limits (
                user_id     INTEGER NOT NULL,
                hour_bucket TEXT    NOT NULL,
                count       INTEGER DEFAULT 0,
                PRIMARY KEY (user_id, hour_bucket)
            );
            CREATE TABLE IF NOT EXISTS stats (
                key   TEXT PRIMARY KEY,
                value TEXT
            );
            CREATE TABLE IF NOT EXISTS blacklist (
                username TEXT PRIMARY KEY,
                reason   TEXT,
                ts       TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS user_settings (
                user_id     INTEGER PRIMARY KEY,
                lang        TEXT    DEFAULT 'he',
                quiet       INTEGER DEFAULT 0,
                min_members INTEGER DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS accepted_terms (
                user_id INTEGER PRIMARY KEY,
                ts      TEXT DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS growth (
                username TEXT    NOT NULL,
                members  INTEGER NOT NULL,
                ts       TEXT    DEFAULT (datetime('now'))
            );
            CREATE TABLE IF NOT EXISTS alerts (
                user_id   INTEGER NOT NULL,
                keyword   TEXT    NOT NULL,
                last_json TEXT    DEFAULT '[]',
                ts        TEXT    DEFAULT (datetime('now')),
                PRIMARY KEY (user_id, keyword)
            );
            CREATE INDEX IF NOT EXISTS idx_growth_username ON growth(username);
            CREATE INDEX IF NOT EXISTS idx_growth_ts       ON growth(ts);
            CREATE INDEX IF NOT EXISTS idx_searches_user   ON searches(user_id);
        """)
        await con.commit()
    log.info("DB initialized ✅")

async def db_log_search(user_id, query, count):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT INTO searches (user_id,query,results_cnt) VALUES (?,?,?)",
            (user_id, query, count)
        )
        await con.execute(
            "INSERT INTO stats (key,value) VALUES ('total_searches','1') "
            "ON CONFLICT(key) DO UPDATE SET value=CAST(value AS INTEGER)+1"
        )
        await con.commit()

async def db_get_history(user_id, limit=10) -> list[dict]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute(
            "SELECT query,results_cnt,ts FROM searches WHERE user_id=? ORDER BY id DESC LIMIT ?",
            (user_id, limit)
        )
        rows = await cur.fetchall()
    return [{'query': r[0], 'count': r[1], 'ts': r[2]} for r in rows]

async def db_check_rate(user_id) -> bool:
    bucket = datetime.utcnow().strftime('%Y-%m-%d-%H')
    async with aiosqlite.connect(DB_PATH) as con:
        cur   = await con.execute(
            "SELECT count FROM rate_limits WHERE user_id=? AND hour_bucket=?",
            (user_id, bucket)
        )
        row   = await cur.fetchone()
        count = row[0] if row else 0
        if count >= RATE_LIMIT_PER_HOUR:
            return False
        await con.execute(
            "INSERT INTO rate_limits (user_id,hour_bucket,count) VALUES (?,?,1) "
            "ON CONFLICT(user_id,hour_bucket) DO UPDATE SET count=count+1",
            (user_id, bucket)
        )
        await con.commit()
    return True

async def db_get_cache(query) -> list | None:
    # Try Redis first
    r = await _get_redis()
    if r:
        try:
            data = await r.get(f"cache:{query}")
            if data:
                return json.loads(data)
        except Exception:
            pass
    # Fallback to SQLite
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute(
            "SELECT data FROM cache WHERE query=? AND expires_at > datetime('now')",
            (query,)
        )
        row = await cur.fetchone()
    return json.loads(row[0]) if row else None

async def db_set_cache(query, data):
    exp = (datetime.utcnow() + timedelta(seconds=CACHE_TTL_SECONDS)).isoformat()
    # Try Redis first
    r = await _get_redis()
    if r:
        try:
            await r.setex(f"cache:{query}", CACHE_TTL_SECONDS, json.dumps(data, ensure_ascii=False))
        except Exception:
            pass
    # Always write to SQLite as backup
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT OR REPLACE INTO cache (query,data,expires_at) VALUES (?,?,?)",
            (query, json.dumps(data, ensure_ascii=False), exp)
        )
        await con.commit()

async def db_clear_cache(query=''):
    r = await _get_redis()
    if r:
        try:
            if query:
                await r.delete(f"cache:{query}")
            else:
                keys = await r.keys("cache:*")
                if keys:
                    await r.delete(*keys)
        except Exception:
            pass
    async with aiosqlite.connect(DB_PATH) as con:
        if query:
            await con.execute("DELETE FROM cache WHERE query=?", (query,))
        else:
            await con.execute("DELETE FROM cache")
        await con.commit()

async def db_subscribe(user_id, query, count):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT OR REPLACE INTO subscriptions (user_id,query,last_count) VALUES (?,?,?)",
            (user_id, query, count)
        )
        await con.commit()

async def db_unsubscribe(user_id, query):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "DELETE FROM subscriptions WHERE user_id=? AND query=?",
            (user_id, query)
        )
        await con.commit()

async def db_get_subscriptions() -> list[dict]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute("SELECT user_id,query,last_count FROM subscriptions")
        rows = await cur.fetchall()
    return [{'user_id': r[0], 'query': r[1], 'last_count': r[2]} for r in rows]

async def db_get_stats() -> dict:
    async with aiosqlite.connect(DB_PATH) as con:
        cur   = await con.execute("SELECT key,value FROM stats")
        stats = {r[0]: r[1] for r in await cur.fetchall()}
        cur   = await con.execute(
            "SELECT query,COUNT(*) c FROM searches GROUP BY query ORDER BY c DESC LIMIT 5"
        )
        top   = await cur.fetchall()
        cur   = await con.execute("SELECT COUNT(DISTINCT user_id) FROM searches")
        users = (await cur.fetchone())[0]
        cur   = await con.execute("SELECT COUNT(*) FROM blacklist")
        bl    = (await cur.fetchone())[0]
        cur   = await con.execute("SELECT COUNT(*) FROM subscriptions")
        subs  = (await cur.fetchone())[0]
        cur   = await con.execute("SELECT COUNT(*) FROM alerts")
        alrts = (await cur.fetchone())[0]
    stats['top_queries']   = top
    stats['unique_users']  = users
    stats['blacklisted']   = bl
    stats['subscriptions'] = subs
    stats['alerts']        = alrts
    return stats

async def db_is_blacklisted(username) -> bool:
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute("SELECT 1 FROM blacklist WHERE username=?", (username.lower(),))
        row = await cur.fetchone()
    return row is not None

async def db_add_blacklist(username, reason=''):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT OR IGNORE INTO blacklist (username,reason) VALUES (?,?)",
            (username.lower(), reason)
        )
        await con.commit()

async def db_get_user_settings(user_id) -> dict:
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute(
            "SELECT lang,quiet,min_members FROM user_settings WHERE user_id=?",
            (user_id,)
        )
        row = await cur.fetchone()
    if row:
        return {'lang': row[0], 'quiet': bool(row[1]), 'min_members': row[2]}
    return {'lang': 'he', 'quiet': False, 'min_members': 0}

async def db_set_user_setting(user_id, key, value):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            f"INSERT INTO user_settings (user_id,{key}) VALUES (?,?) "
            f"ON CONFLICT(user_id) DO UPDATE SET {key}=?",
            (user_id, value, value)
        )
        await con.commit()

async def db_accepted_terms(user_id) -> bool:
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute("SELECT 1 FROM accepted_terms WHERE user_id=?", (user_id,))
        row = await cur.fetchone()
    return row is not None

async def db_accept_terms(user_id):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute("INSERT OR IGNORE INTO accepted_terms (user_id) VALUES (?)", (user_id,))
        await con.commit()

async def db_save_growth(username: str, members: int):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT INTO growth (username,members) VALUES (?,?)",
            (username.lower(), members)
        )
        await con.commit()

async def db_get_growth(username: str, hours: int = 168) -> list[dict]:
    since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute(
            "SELECT members,ts FROM growth WHERE username=? AND ts>? ORDER BY ts",
            (username.lower(), since)
        )
        rows = await cur.fetchall()
    return [{'members': r[0], 'ts': r[1]} for r in rows]

async def db_get_trending(hours: int = 24, limit: int = 10) -> list[dict]:
    since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute("""
            SELECT username,
                   MAX(members) - MIN(members) AS growth,
                   MAX(members) AS current_val,
                   MIN(members) AS start_val
            FROM growth
            WHERE ts > ?
            GROUP BY username
            HAVING COUNT(*) >= 2 AND growth > 0
            ORDER BY growth DESC
            LIMIT ?
        """, (since, limit))
        rows = await cur.fetchall()
    return [{'username': r[0], 'growth': r[1], 'current': r[2], 'start': r[3]} for r in rows]

async def db_get_all_user_ids() -> list[int]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute("SELECT DISTINCT user_id FROM searches")
        rows = await cur.fetchall()
    return [r[0] for r in rows]

# ── Alerts DB ──
async def db_add_alert(user_id: int, keyword: str):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "INSERT OR IGNORE INTO alerts (user_id,keyword) VALUES (?,?)",
            (user_id, keyword.lower())
        )
        await con.commit()

async def db_remove_alert(user_id: int, keyword: str):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "DELETE FROM alerts WHERE user_id=? AND keyword=?",
            (user_id, keyword.lower())
        )
        await con.commit()

async def db_get_user_alerts(user_id: int) -> list[str]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute("SELECT keyword FROM alerts WHERE user_id=?", (user_id,))
        rows = await cur.fetchall()
    return [r[0] for r in rows]

async def db_get_all_alerts() -> list[dict]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur  = await con.execute("SELECT user_id,keyword,last_json FROM alerts")
        rows = await cur.fetchall()
    return [{'user_id': r[0], 'keyword': r[1], 'last_found': json.loads(r[2] or '[]')} for r in rows]

async def db_update_alert_found(user_id: int, keyword: str, found: list):
    async with aiosqlite.connect(DB_PATH) as con:
        await con.execute(
            "UPDATE alerts SET last_json=? WHERE user_id=? AND keyword=?",
            (json.dumps(found, ensure_ascii=False), user_id, keyword)
        )
        await con.commit()

# ── Rankings DB ──
async def db_get_top_users(limit: int = 10) -> list[dict]:
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute(
            "SELECT user_id, COUNT(*) c, SUM(results_cnt) total "
            "FROM searches GROUP BY user_id ORDER BY c DESC LIMIT ?",
            (limit,)
        )
        rows = await cur.fetchall()
    return [{'user_id': r[0], 'searches': r[1], 'total_results': r[2] or 0} for r in rows]

async def db_get_hourly_stats(hours: int = 24) -> list[dict]:
    since = (datetime.utcnow() - timedelta(hours=hours)).isoformat()
    async with aiosqlite.connect(DB_PATH) as con:
        cur = await con.execute(
            "SELECT strftime('%H', ts) h, COUNT(*) c "
            "FROM searches WHERE ts > ? GROUP BY h ORDER BY h",
            (since,)
        )
        rows = await cur.fetchall()
    return [{'hour': r[0], 'count': r[1]} for r in rows]

# ─────────────────────────────────────────────
#  עזרי שפה, ספאם, טיפוס
# ─────────────────────────────────────────────

def detect_language(text: str) -> str:
    hebrew = len(re.findall(r'[֐-׿]', text))
    arabic = len(re.findall(r'[؀-ۿ]', text))
    cyril  = len(re.findall(r'[Ѐ-ӿ]', text))
    latin  = len(re.findall(r'[a-zA-Z]', text))
    if hebrew > arabic and hebrew > cyril and hebrew > latin: return 'he'
    if arabic > hebrew and arabic > cyril:                    return 'ar'
    if cyril  > latin:                                        return 'ru'
    return 'en'

def expand_query(query: str) -> list[str]:
    queries = [query]
    q_lower = query.lower()
    for key, syns in SYNONYMS.items():
        if key in q_lower or q_lower in key:
            queries += syns
    return list(set(queries))[:6]

def is_spam_channel(title: str, username: str) -> bool:
    combined = (title + ' ' + username).lower()
    return any(kw.lower() in combined for kw in SPAM_KEYWORDS)

def levenshtein(a: str, b: str) -> int:
    if len(a) < len(b):
        a, b = b, a
    if not b:
        return len(a)
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a):
        curr = [i + 1]
        for j, cb in enumerate(b):
            curr.append(min(prev[j + 1] + 1, curr[j] + 1, prev[j] + (ca != cb)))
        prev = curr
    return prev[-1]

TYPO_DICT: dict[str, str] = {
    'ביטקיון': 'ביטקוין',
    'ביטקוון': 'ביטקוין',
    'קריפטא':  'קריפטו',
}

def fix_typo(query: str) -> str:
    q = query.strip()
    if q in TYPO_DICT:
        return TYPO_DICT[q]
    for known, correct in TYPO_DICT.items():
        if levenshtein(q, known) <= 1:
            return correct
    return q

async def translate_query(session: aiohttp.ClientSession, query: str) -> str | None:
    lang = detect_language(query)
    if lang not in ('he', 'ar'):
        return None
    try:
        url = (
            f"https://api.mymemory.translated.net/get"
            f"?q={urllib.parse.quote(query)}&langpair={lang}|en"
        )
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=6)) as r:
            data       = await r.json(content_type=None)
            translated = data.get('responseData', {}).get('translatedText', '')
            if translated and translated.lower() != query.lower():
                return translated
    except Exception as e:
        log.debug(f"Translation error: {e}")
    return None

# ─────────────────────────────────────────────
#  ניקוד רלוונטיות + Health Score + קטגוריה
# ─────────────────────────────────────────────

def _relevance_score(r: dict, query: str) -> float:
    score = 0.0
    q     = query.lower()
    words = q.split()
    title    = r.get('title', '').lower()
    username = r.get('username', '').lower()
    about    = r.get('about', '').lower()

    if q in title:          score += 10.0
    if q in username:       score += 6.0
    if q in about:          score += 2.0
    for w in words:
        if w in title:      score += 3.0
        if w in username:   score += 2.0
        if w in about:      score += 0.5
    return score

def _health_score(r: dict) -> int:
    score    = 0
    members  = r.get('members', 0)
    msgs_24h = r.get('msgs_24h', 0)
    activity = r.get('activity', '')

    if members >= 100_000: score += 40
    elif members >= 10_000: score += 30
    elif members >= 1_000:  score += 20
    elif members >= 100:    score += 10
    else:                   score += 3

    if '🟢' in activity:   score += 40
    elif '🟡' in activity: score += 25
    elif '🟠' in activity: score += 10

    if msgs_24h >= 20:   score += 20
    elif msgs_24h >= 10: score += 15
    elif msgs_24h >= 5:  score += 10
    elif msgs_24h >= 1:  score += 5

    return min(score, 100)

def _health_bar(score: int) -> str:
    filled = score // 10
    bar    = '█' * filled + '░' * (10 - filled)
    emoji  = '🟢' if score >= 70 else '🟡' if score >= 40 else '🔴'
    return f"{emoji} {bar} {score}/100"

def auto_categorize(title: str, about: str) -> str:
    combined = (title + ' ' + about).lower()
    best_cat, best_cnt = 'כללי', 0
    for cat, keywords in CATEGORIES.items():
        cnt = sum(1 for kw in keywords if kw in combined)
        if cnt > best_cnt:
            best_cnt, best_cat = cnt, cat
    return best_cat

# ─────────────────────────────────────────────
#  חיפוש חיצוני — async (aiohttp + proxy)
# ─────────────────────────────────────────────

def _extract_tme(html: str) -> set[str]:
    found = set()
    for m in re.finditer(r'https?://t\.me/([a-zA-Z0-9_]{5,})', html):
        full = m.group(0).rstrip('/')
        if not re.search(r'/(share|joinchat|s|addstickers|setlanguage)$', full):
            found.add(full)
    return found

async def _search_engine_async(
    session: aiohttp.ClientSession, url: str, name: str
) -> set[str]:
    proxy = _next_proxy()
    try:
        async with session.get(
            url, headers=HEADERS,
            timeout=aiohttp.ClientTimeout(total=10),
            proxy=proxy
        ) as r:
            html = await r.text()
            return _extract_tme(html)
    except Exception as e:
        log.warning(f"{name} error: {e}")
        return set()

async def search_all_external_async(
    session: aiohttp.ClientSession, queries: list[str]
) -> list[str]:
    tasks = []
    for q in queries:
        for name, url_fn in SEARCH_SOURCES.items():
            try:
                url = url_fn(q)
            except Exception:
                continue
            tasks.append(_search_engine_async(session, url, name))
    sets = await asyncio.gather(*tasks, return_exceptions=True)
    merged: set[str] = set()
    for s in sets:
        if isinstance(s, set):
            merged |= s
    log.info(f"External: {len(merged)} links from {len(SEARCH_SOURCES)} sources × {len(queries)} queries")
    return list(merged)

# ─────────────────────────────────────────────
#  חיפוש פנימי ב-Telegram
# ─────────────────────────────────────────────

async def telegram_search_variants(user_client, queries: list[str]) -> dict[str, dict]:
    results: dict[str, dict] = {}
    chars    = "abcdefghijklmnopqrstuvwxyz0123456789"
    variants = []
    for q in queries:
        variants.append(q)
        variants += [f"{q} {c}" for c in chars[:SEARCH_VARIANTS]]

    async def _one(q):
        try:
            res = await user_client(functions.contacts.SearchRequest(q=q, limit=100))
            for chat in res.chats:
                if hasattr(chat, 'username') and chat.username:
                    results[chat.username] = {
                        'title':    chat.title,
                        'username': chat.username,
                        'members':  getattr(chat, 'participants_count', 0) or 0,
                    }
            await asyncio.sleep(0.2)
        except FloodWaitError as e:
            log.warning(f"FloodWait '{q}': sleeping {e.seconds}s")
            await asyncio.sleep(e.seconds)
        except Exception as e:
            log.debug(f"Variant error '{q}': {e}")

    for i in range(0, len(variants), 10):
        await asyncio.gather(*[_one(v) for v in variants[i:i + 10]])
        await asyncio.sleep(0.5)
    return results

async def get_channel_activity(user_client, entity) -> dict:
    try:
        since = datetime.utcnow() - timedelta(hours=24)
        hist  = await user_client(GetHistoryRequest(
            peer=entity, limit=50, offset_date=None,
            offset_id=0, max_id=0, min_id=0, add_offset=0, hash=0
        ))
        recent = [
            m for m in hist.messages
            if hasattr(m, 'date') and m.date and m.date.replace(tzinfo=None) > since
        ]
        return {
            'msgs_24h': len(recent),
            'last_msg': hist.messages[0].date.replace(tzinfo=None) if hist.messages else None,
        }
    except Exception:
        return {'msgs_24h': 0, 'last_msg': None}

async def get_recent_messages(user_client, entity, limit: int = 5) -> list[str]:
    try:
        hist = await user_client(GetHistoryRequest(
            peer=entity, limit=limit, offset_date=None,
            offset_id=0, max_id=0, min_id=0, add_offset=0, hash=0
        ))
        return [
            f"• {m.message[:100].replace(chr(10), ' ')}"
            for m in hist.messages
            if hasattr(m, 'message') and m.message
        ]
    except Exception:
        return []

async def enrich_channel_info(user_client, username: str) -> dict | None:
    sem = ENRICH_SEM or asyncio.Semaphore(5)
    async with sem:
        if await db_is_blacklisted(username):
            return None
        try:
            entity = await user_client(ResolveUsernameRequest(username))
            if not hasattr(entity, 'chats') or not entity.chats:
                return None
            chat    = entity.chats[0]
            full    = await user_client(GetFullChannelRequest(chat))
            members = getattr(full.full_chat, 'participants_count', 0) or 0
            about   = getattr(full.full_chat, 'about', '') or ''
            title   = chat.title

            if is_spam_channel(title, username):
                await db_add_blacklist(username, 'auto-spam-filter')
                return None

            lang_code = detect_language(title + ' ' + about)
            lang_flag = LANGUAGE_MAP.get(lang_code, '🌐')
            activity  = await get_channel_activity(user_client, chat)
            msgs_24h  = activity['msgs_24h']
            last_msg  = activity['last_msg']

            if msgs_24h >= 10:    activity_label = "🟢 פעיל מאוד"
            elif msgs_24h >= 1:   activity_label = "🟡 פעיל"
            elif last_msg and (datetime.utcnow() - last_msg).days < 7:
                                  activity_label = "🟠 נמוך"
            else:                 activity_label = "💀 לא פעיל"

            category = auto_categorize(title, about)
            asyncio.create_task(db_save_growth(username, members))

            result = {
                'title':     title,
                'username':  username,
                'members':   members,
                'about':     about[:120],
                'link':      f"https://t.me/{username}",
                'lang':      lang_flag,
                'lang_code': lang_code,
                'activity':  activity_label,
                'msgs_24h':  msgs_24h,
                'category':  category,
            }
            result['health'] = _health_score(result)
            return result
        except FloodWaitError as e:
            log.warning(f"FloodWait enrich {username}: sleeping {e.seconds}s")
            await asyncio.sleep(e.seconds)
            return None
        except Exception as e:
            log.debug(f"Enrich error {username}: {e}")
            return None

# ─────────────────────────────────────────────
#  חיפוש מאוחד
# ─────────────────────────────────────────────

active_searches: set[int] = set()

async def full_search(
    user_client,
    query:        str,
    min_members:  int  = 0,
    force_fresh:  bool = False,
    lang_filter:  str  = '',
    cat_filter:   str  = '',
    progress_cb        = None,
) -> list[dict]:
    query = fix_typo(query)

    if not force_fresh:
        cached = await db_get_cache(query)
        if cached:
            log.info(f"Cache hit: '{query}'")
            filtered = [r for r in cached if r['members'] >= min_members]
            if lang_filter:
                filtered = [r for r in filtered if r.get('lang_code') == lang_filter]
            if cat_filter:
                filtered = [r for r in filtered if r.get('category') == cat_filter]
            return filtered

    log.info(f"Full search: '{query}'")
    if progress_cb:
        await progress_cb("🔍 שלב 1/4: מרחיב שאילתה ומתרגם...")

    async with aiohttp.ClientSession() as session:
        queries    = expand_query(query)
        translated = await translate_query(session, query)
        if translated:
            queries.append(translated)
            queries += expand_query(translated)
        queries = list(set(queries))[:8]

        if progress_cb:
            await progress_cb("🌐 שלב 2/4: חיפוש חיצוני (30 מקורות)...")

        external_links, internal_results = await asyncio.gather(
            search_all_external_async(session, queries),
            telegram_search_variants(user_client, queries)
        )

    if progress_cb:
        await progress_cb("📡 שלב 3/4: מאחד תוצאות...")

    all_usernames: set[str] = set(internal_results.keys())
    for link in external_links:
        m = re.search(r't\.me/([a-zA-Z0-9_]{5,})', link)
        if m:
            all_usernames.add(m.group(1))

    if progress_cb:
        await progress_cb(f"✨ שלב 4/4: מעשיר {len(all_usernames)} ערוצים...")

    results = await asyncio.gather(*[enrich_channel_info(user_client, u) for u in all_usernames])
    enriched = [
        r for r in results
        if r
        and r['members'] >= min_members
        and (not lang_filter or r.get('lang_code') == lang_filter)
        and (not cat_filter  or r.get('category')  == cat_filter)
    ]

    # מיון: רלוונטיות × health score × מנויים
    enriched.sort(
        key=lambda x: (
            _relevance_score(x, query) * 2 +
            x.get('health', 0) * 0.5 +
            min(x['members'] / 10000, 10)
        ),
        reverse=True
    )
    await db_set_cache(query, enriched)
    return enriched

# ─────────────────────────────────────────────
#  פורמט ותוצאות
# ─────────────────────────────────────────────

def format_result(r: dict, rank: int, show_health: bool = True) -> str:
    members_str = f"{r['members']:,}" if r['members'] else "?"
    about       = f"\n📝 _{r['about']}_" if r.get('about') else ""
    health_line = f"\n{_health_bar(r.get('health', 0))}" if show_health else ""
    cat         = f" [{r.get('category','כללי')}]" if r.get('category') else ""
    return (
        f"**{rank}. {r['title']}** {r.get('lang', '🌐')}{cat}\n"
        f"{about}"
        f"{health_line}\n"
        f"👥 {members_str} | {r.get('activity', '')} ({r.get('msgs_24h', 0)} הודעות/יום)\n"
        f"🔗 {r['link']}"
    )

def format_compare(r1: dict, r2: dict) -> str:
    def row(label, v1, v2):
        return f"**{label}**\n  A: {v1}\n  B: {v2}\n"

    winner_members = "A ✅" if r1['members'] > r2['members'] else "B ✅"
    winner_health  = "A ✅" if r1.get('health',0) > r2.get('health',0) else "B ✅"
    winner_msgs    = "A ✅" if r1.get('msgs_24h',0) > r2.get('msgs_24h',0) else "B ✅"

    return (
        f"⚖️ **השוואת ערוצים**\n\n"
        f"🅰️ **{r1['title']}** (@{r1['username']})\n"
        f"🅱️ **{r2['title']}** (@{r2['username']})\n\n"
        f"👥 מנויים ({winner_members})\n"
        f"  A: {r1['members']:,}  |  B: {r2['members']:,}\n\n"
        f"❤️ Health Score ({winner_health})\n"
        f"  A: {_health_bar(r1.get('health',0))}\n"
        f"  B: {_health_bar(r2.get('health',0))}\n\n"
        f"📨 הודעות/יום ({winner_msgs})\n"
        f"  A: {r1.get('msgs_24h',0)}  |  B: {r2.get('msgs_24h',0)}\n\n"
        f"🏷 קטגוריה\n"
        f"  A: {r1.get('category','כללי')}  |  B: {r2.get('category','כללי')}\n\n"
        f"🔗 A: {r1['link']}\n"
        f"🔗 B: {r2['link']}"
    )

def results_to_csv(results: list[dict]) -> bytes:
    out    = io.StringIO()
    fields = ['rank', 'title', 'username', 'members', 'health', 'category',
              'about', 'link', 'lang', 'activity', 'msgs_24h']
    w = csv.DictWriter(out, fieldnames=fields)
    w.writeheader()
    for i, r in enumerate(results, 1):
        row = {k: r.get(k, '') for k in fields}
        row['rank'] = i
        w.writerow(row)
    return out.getvalue().encode('utf-8-sig')

def results_to_xlsx(results: list[dict]) -> bytes:
    if not HAS_OPENPYXL:
        return results_to_csv(results)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "תוצאות חיפוש"

    headers = ['#', 'שם', 'Username', 'מנויים', 'Health', 'קטגוריה',
               'תיאור', 'קישור', 'שפה', 'פעילות', 'הודעות/יום']
    ws.append(headers)

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="2B5CE6")
    hdr_align = Alignment(horizontal='center')
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    for i, r in enumerate(results, 1):
        health = r.get('health', 0)
        row_data = [
            i, r.get('title',''), r.get('username',''), r.get('members',0),
            health, r.get('category',''), r.get('about',''), r.get('link',''),
            r.get('lang',''), r.get('activity',''), r.get('msgs_24h',0),
        ]
        ws.append(row_data)
        # צבע שורה לפי health
        fill_color = "C8E6C9" if health >= 70 else "FFF9C4" if health >= 40 else "FFCDD2"
        for cell in ws[ws.max_row]:
            cell.fill = PatternFill("solid", fgColor=fill_color)

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def results_to_pdf(results: list[dict], query: str) -> bytes:
    if not HAS_FPDF:
        return results_to_csv(results)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, f"Telegram Search: {query}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | {len(results)} results", ln=True, align="C")
    pdf.ln(4)

    # Table header
    pdf.set_fill_color(43, 92, 230)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    col_w = [8, 50, 30, 20, 15, 25, 42]
    headers_en = ["#", "Title", "Username", "Members", "Health", "Category", "Link"]
    for w, h in zip(col_w, headers_en):
        pdf.cell(w, 7, h, border=1, fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for i, r in enumerate(results, 1):
        health = r.get('health', 0)
        if health >= 70:   pdf.set_fill_color(200, 230, 201)
        elif health >= 40: pdf.set_fill_color(255, 249, 196)
        else:              pdf.set_fill_color(255, 205, 210)
        fill = True

        title_ascii = r.get('title','')[:28].encode('ascii','replace').decode()
        user        = r.get('username','')[:18]
        members     = f"{r.get('members',0):,}"
        cat         = r.get('category','')[:12].encode('ascii','replace').decode()
        link        = f"t.me/{user}"

        row_vals = [str(i), title_ascii, user, members, str(health), cat, link]
        for w, v in zip(col_w, row_vals):
            pdf.cell(w, 6, v, border=1, fill=fill)
        pdf.ln()

    return pdf.output()

def draw_growth_chart(growth_data: list[dict]) -> str:
    if len(growth_data) < 2:
        return "אין מספיק נתונים"
    members = [d['members'] for d in growth_data]
    mn, mx  = min(members), max(members)
    if mx == mn:
        return "ללא שינוי"
    HEIGHT = 8
    rows = []
    for d in growth_data[-12:]:
        pct  = (d['members'] - mn) / (mx - mn)
        bars = int(pct * HEIGHT)
        rows.append(f"{d['members']:>8,} │{'█' * bars}{'░' * (HEIGHT - bars)}")
    return f"📈 {mn:,} → {mx:,}  (+{mx - mn:,})\n" + "\n".join(rows)

# ─────────────────────────────────────────────
#  Inline Buttons
# ─────────────────────────────────────────────

def make_buttons(query: str, page: int, total: int) -> list:
    btns = []
    nav  = []
    if page > 0:
        nav.append(Button.inline(f"⬅️ {page}", data=f"page:{query}:{page - 1}"))
    if (page + 1) * 10 < total:
        nav.append(Button.inline(f"{page + 2} ➡️", data=f"page:{query}:{page + 1}"))
    if nav:
        btns.append(nav)
    btns.append([
        Button.inline("🔄 רענן",   data=f"refresh:{query}"),
        Button.inline("🔔 הרשם",   data=f"sub:{query}"),
        Button.inline("📤 שתף",    data=f"share:{query}"),
    ])
    btns.append([
        Button.inline("📊 Excel",  data=f"xlsx:{query}"),
        Button.inline("📥 CSV",    data=f"csv:{query}"),
        Button.inline("📄 PDF",    data=f"pdf:{query}"),
    ])
    return btns

# ─────────────────────────────────────────────
#  סטטוס ו-Dashboard
# ─────────────────────────────────────────────

async def get_status_text() -> str:
    uptime  = datetime.utcnow() - BOT_START_TIME
    hours   = int(uptime.total_seconds() // 3600)
    minutes = int((uptime.total_seconds() % 3600) // 60)
    try:
        mem_mb = psutil.Process().memory_info().rss / 1024 / 1024
        cpu    = psutil.cpu_percent(interval=0.5)
    except Exception:
        mem_mb, cpu = 0, 0
    db_size = os.path.getsize(DB_PATH) / 1024 if os.path.exists(DB_PATH) else 0
    redis_ok = "✅" if await _get_redis() else "❌ (SQLite)"
    s = await db_get_stats()
    return (
        f"📡 **סטטוס הבוט v3.0**\n\n"
        f"⏱ Uptime: **{hours}h {minutes}m**\n"
        f"💾 RAM: **{mem_mb:.1f} MB** | ⚙️ CPU: **{cpu:.1f}%**\n"
        f"🗄 DB: **{db_size:.1f} KB** | Redis: **{redis_ok}**\n"
        f"🔍 חיפושים: **{s.get('total_searches', 0)}**\n"
        f"👤 משתמשים: **{s.get('unique_users', 0)}**\n"
        f"🔔 מנויים: **{s.get('subscriptions', 0)}** | "
        f"🔔 Alerts: **{s.get('alerts', 0)}**\n"
        f"🚫 Blacklist: **{s.get('blacklisted', 0)}**\n"
        f"🐍 Python: **{platform.python_version()}**\n"
        f"📊 Excel: **{'✅' if HAS_OPENPYXL else '❌'}** | "
        f"📄 PDF: **{'✅' if HAS_FPDF else '❌ pip install fpdf2'}**\n"
        f"🔀 Proxies: **{len(PROXY_LIST)}**"
    )

async def get_dashboard_text() -> str:
    hourly  = await db_get_hourly_stats(24)
    s       = await db_get_stats()
    top     = await db_get_top_users(5)
    trending = await db_get_trending(24, 3)

    # גרף שעתי ASCII
    if hourly:
        mx  = max(h['count'] for h in hourly) or 1
        bar = ""
        for h in hourly:
            filled = int(h['count'] / mx * 10)
            bar += f"  {h['hour']}:00 {'█' * filled}{'░' * (10 - filled)} {h['count']}\n"
    else:
        bar = "  אין נתונים עדיין\n"

    top_str = "\n".join(
        f"  {i+1}. uid={u['user_id']} — {u['searches']} חיפושים, {u['total_results']} תוצאות"
        for i, u in enumerate(top)
    ) or "  אין עדיין"

    trend_str = "\n".join(
        f"  @{t['username']} +{t['growth']:,}"
        for t in trending
    ) or "  אין עדיין"

    top_q = "\n".join(f"  • `{q[0]}` ({q[1]}x)" for q in s.get('top_queries', [])) or "  אין עדיין"

    return (
        f"📊 **Dashboard v3.0**\n\n"
        f"**📈 חיפושים לפי שעה (24h):**\n```\n{bar}```\n"
        f"**🏆 Top 5 משתמשים:**\n{top_str}\n\n"
        f"**🔥 Trending (24h):**\n{trend_str}\n\n"
        f"**🔍 Top Queries:**\n{top_q}"
    )

# ─────────────────────────────────────────────
#  רישום handlers
# ─────────────────────────────────────────────

def register_handlers(bot_client, user_client):

    pending_search: dict[int, dict] = {}

    async def _check_terms(event) -> bool:
        if await db_accepted_terms(event.sender_id):
            return True
        await event.respond(
            "📋 **תנאי שימוש**\n\n"
            "הבוט מחפש קבוצות וערוצים בטלגרם.\n"
            "• אין להשתמש לפעילות בלתי חוקית\n"
            "• הנתונים נשמרים לצורך שיפור השירות\n\n"
            "לחץ **אישור** להמשך:",
            buttons=[[Button.inline("✅ אני מסכים/ה", data="accept_terms")]]
        )
        return False

    @bot_client.on(events.CallbackQuery(data=b"accept_terms"))
    async def cb_accept_terms(event):
        await db_accept_terms(event.sender_id)
        await event.edit("✅ **תודה! כעת תוכל להשתמש בבוט.**\nשלח מילת חיפוש כלשהי.")

    # ── /start ──
    @bot_client.on(events.NewMessage(pattern='/start'))
    async def cmd_start(event):
        await event.respond(
            "👋 **ברוך הבא לבוט החיפוש v3.0!**\n\n"
            "שלח מילת חיפוש חופשית, או השתמש בפקודות:\n\n"
            "🔍 `/search <שאילתה>` — חיפוש רגיל\n"
            "🚀 `/deep <שאילתה>` — חיפוש עמוק (ללא מטמון)\n"
            "🧠 `/smart` — חיפוש אינטרקטיבי עם פילטרים\n"
            "📊 `/filter <מנויים> <שאילתה>` — סף מנויים\n"
            "🏷 `/category <שם>` — חיפוש לפי קטגוריה\n"
            "⚖️ `/compare @chan1 @chan2` — השוואת שני ערוצים\n"
            "🔗 `/similar @channel` — ערוצים דומים\n"
            "👁 `/preview @channel` — תצוגה מקדימה\n"
            "📈 `/growth @channel` — גרף צמיחה\n"
            "🔥 `/trending` — ערוצים עולים ב-24 שעות\n"
            "🔔 `/alert <מילה>` — התראה על ערוץ חדש\n"
            "📋 `/alerts` — רשימת ההתראות שלי\n"
            "❌ `/unalert <מילה>` — מחק התראה\n"
            "📜 `/history` — היסטוריית חיפושים\n"
            "🔔 `/subscribe <שאילתה>` — מנוי לעדכונים\n"
            "🔕 `/quiet` — מצב שקט (Excel בלבד)\n"
            "🌐 `/lang <he/en/ru>` — שפת ממשק\n"
            "⛔ `/cancel` — ביטול חיפוש פעיל\n"
            "📡 `/status` — מצב הבוט\n"
            "🏆 `/top` — דירוג משתמשים\n"
            "🔎 `/sources` — כל מקורות החיפוש\n"
            "🚫 `/blacklist @user` — חסום ערוץ (אדמין)\n"
            "📢 `/broadcast <msg>` — שלח לכולם (אדמין)\n"
            "📊 `/dashboard` — לוח מחוונים (אדמין)\n"
            "📊 `/stats` — סטטיסטיקות (אדמין)"
        )

    # ── /help ──
    @bot_client.on(events.NewMessage(pattern='/help'))
    async def cmd_help(event):
        cats = " · ".join(f"`{c}`" for c in CATEGORIES.keys())
        await event.respond(
            "📖 **מדריך שימוש מלא v3.0**\n\n"
            "**⭐ חדש ב-v3.0:**\n"
            "• **Relevance Score** — דירוג לפי התאמה לשאילתה\n"
            "• **Health Score (0-100)** — ציון בריאות ערוץ משוקלל\n"
            "• **⚖️ Compare** — השוואת שני ערוצים\n"
            "• **🔔 Alerts** — התראה כשערוץ חדש עם מילה מסוימת מופיע\n"
            "• **🏷 Categories** — סינון לפי קטגוריה\n"
            "• **📄 PDF Export** — ייצוא לפורמט PDF\n"
            "• **📊 Dashboard** — גרפי שימוש\n"
            "• **🏆 Top Users** — דירוג משתמשים\n"
            "• **Proxy Pool** — רוטציה בין פרוקסים\n"
            "• **Redis Cache** — מטמון מהיר (אופציונלי)\n\n"
            f"**🏷 קטגוריות זמינות:**\n{cats}\n\n"
            "**ייצוא:** 📊 Excel מעוצב | 📥 CSV | 📄 PDF\n"
            "**Inline Mode:** הקלד `@botname ביטקוין` בכל שיחה"
        )

    # ── /status ──
    @bot_client.on(events.NewMessage(pattern='/status'))
    async def cmd_status(event):
        await event.respond(await get_status_text())

    # ── /dashboard ──
    @bot_client.on(events.NewMessage(pattern='/dashboard'))
    async def cmd_dashboard(event):
        if ADMIN_IDS and event.sender_id not in ADMIN_IDS:
            await event.respond("⛔ פקודת אדמין בלבד.")
            return
        await event.respond(await get_dashboard_text())

    # ── /sources ──
    @bot_client.on(events.NewMessage(pattern='/sources'))
    async def cmd_sources(event):
        categories = {
            "🔎 מנועי חיפוש": ['Google','Bing','DDG','Yandex','Yahoo','Brave','Startpage','Ecosia','Mojeek','Baidu'],
            "📋 קטלוגי טלגרם": ['TGStat','TGStatCom','Telemetr','Combot','Lyzem','HotTG','TgPw',
                                 'Tchannels','TGrам','TelegramDB','TgDev','TgList','TeleList',
                                 'TelegramGroup','TGChannels','TGDir','Tlgrm','SocialBlade'],
            "🌐 מקורות נוספים": ['Reddit','GitHub'],
        }
        lines = [f"📡 **{len(SEARCH_SOURCES)} מקורות חיפוש פעילים:**\n"]
        for cat, names in categories.items():
            lines.append(f"\n**{cat}:**")
            lines.append("  " + " · ".join(f"`{n}`" for n in names))
        if PROXY_LIST:
            lines.append(f"\n🔀 **Proxy Pool:** {len(PROXY_LIST)} פרוקסים פעילים")
        await event.respond("\n".join(lines))

    # ── /top ──
    @bot_client.on(events.NewMessage(pattern='/top'))
    async def cmd_top(event):
        top = await db_get_top_users(10)
        if not top:
            await event.respond("📭 אין נתונים עדיין.")
            return
        lines = [
            f"{'🥇' if i==0 else '🥈' if i==1 else '🥉' if i==2 else f'{i+1}.'} "
            f"user `{u['user_id']}` — **{u['searches']}** חיפושים | {u['total_results']:,} תוצאות"
            for i, u in enumerate(top)
        ]
        await event.respond("🏆 **Top 10 משתמשים:**\n\n" + "\n".join(lines))

    # ── /compare ──
    @bot_client.on(events.NewMessage(pattern=r'/compare @?(\w+)\s+@?(\w+)'))
    async def cmd_compare(event):
        u1 = event.pattern_match.group(1)
        u2 = event.pattern_match.group(2)
        msg = await event.respond(f"⚖️ משווה @{u1} ↔ @{u2}...")
        info1, info2 = await asyncio.gather(
            enrich_channel_info(user_client, u1),
            enrich_channel_info(user_client, u2)
        )
        if not info1:
            await msg.edit(f"❌ לא נמצא @{u1}")
            return
        if not info2:
            await msg.edit(f"❌ לא נמצא @{u2}")
            return
        await msg.edit(format_compare(info1, info2))

    # ── /alert ──
    @bot_client.on(events.NewMessage(pattern=r'/alert (.+)'))
    async def cmd_alert(event):
        keyword = event.pattern_match.group(1).strip()
        existing = await db_get_user_alerts(event.sender_id)
        if len(existing) >= 10:
            await event.respond("⚠️ מקסימום 10 התראות. מחק אחת קודם עם `/unalert`.")
            return
        await db_add_alert(event.sender_id, keyword)
        await event.respond(
            f"🔔 **התראה נוספה:** `{keyword}`\n"
            f"תקבל הודעה כשערוצים חדשים עם מילה זו יתגלו."
        )

    # ── /alerts ──
    @bot_client.on(events.NewMessage(pattern='/alerts'))
    async def cmd_alerts(event):
        kws = await db_get_user_alerts(event.sender_id)
        if not kws:
            await event.respond("📭 אין לך התראות פעילות.\nהוסף עם `/alert <מילה>`")
            return
        lines = [f"  {i+1}. `{kw}`" for i, kw in enumerate(kws)]
        await event.respond(
            f"🔔 **ההתראות שלך ({len(kws)}/10):**\n\n" + "\n".join(lines) +
            "\n\nמחק עם `/unalert <מילה>`"
        )

    # ── /unalert ──
    @bot_client.on(events.NewMessage(pattern=r'/unalert (.+)'))
    async def cmd_unalert(event):
        keyword = event.pattern_match.group(1).strip()
        await db_remove_alert(event.sender_id, keyword)
        await event.respond(f"✅ התראה הוסרה: `{keyword}`")

    # ── /category ──
    @bot_client.on(events.NewMessage(pattern=r'/category (.+)'))
    async def cmd_category(event):
        cat = event.pattern_match.group(1).strip()
        if cat not in CATEGORIES:
            cats = " | ".join(CATEGORIES.keys())
            await event.respond(f"❌ קטגוריה לא קיימת.\nזמינות: {cats}")
            return
        keywords = CATEGORIES[cat]
        query    = keywords[0]
        await _do_search(event, query, cat_filter=cat,
                         custom_title=f"קטגוריה: {cat}")

    # ── /history ──
    @bot_client.on(events.NewMessage(pattern='/history'))
    async def cmd_history(event):
        hist = await db_get_history(event.sender_id)
        if not hist:
            await event.respond("📭 אין היסטוריית חיפושים עדיין.")
            return
        lines = [
            f"{i + 1}. `{h['query']}` — {h['count']} תוצאות ({h['ts'][:16]})"
            for i, h in enumerate(hist)
        ]
        await event.respond("📜 **10 חיפושים אחרונים:**\n\n" + "\n".join(lines))

    # ── /stats ──
    @bot_client.on(events.NewMessage(pattern='/stats'))
    async def cmd_stats(event):
        if ADMIN_IDS and event.sender_id not in ADMIN_IDS:
            await event.respond("⛔ פקודת אדמין בלבד.")
            return
        s   = await db_get_stats()
        top = "\n".join([f"  • `{q[0]}` ({q[1]}x)" for q in s.get('top_queries', [])])
        await event.respond(
            f"📊 **סטטיסטיקות:**\n\n"
            f"🔍 חיפושים: **{s.get('total_searches', 0)}**\n"
            f"👤 משתמשים: **{s.get('unique_users', 0)}**\n"
            f"🔔 מנויים: **{s.get('subscriptions', 0)}**\n"
            f"🔔 Alerts: **{s.get('alerts', 0)}**\n"
            f"🚫 Blacklist: **{s.get('blacklisted', 0)}**\n\n"
            f"🏆 **TOP 5:**\n{top or 'אין עדיין'}"
        )

    # ── /blacklist ──
    @bot_client.on(events.NewMessage(pattern=r'/blacklist (.+)'))
    async def cmd_blacklist(event):
        if ADMIN_IDS and event.sender_id not in ADMIN_IDS:
            await event.respond("⛔ פקודת אדמין בלבד.")
            return
        username = event.pattern_match.group(1).strip().lstrip('@')
        await db_add_blacklist(username, f'manual by admin {event.sender_id}')
        await db_clear_cache('')
        await event.respond(f"🚫 `{username}` נוסף לרשימה השחורה.")

    # ── /broadcast ──
    @bot_client.on(events.NewMessage(pattern=r'/broadcast (.+)'))
    async def cmd_broadcast(event):
        if ADMIN_IDS and event.sender_id not in ADMIN_IDS:
            await event.respond("⛔ פקודת אדמין בלבד.")
            return
        msg      = event.pattern_match.group(1).strip()
        user_ids = await db_get_all_user_ids()
        sent, failed = 0, 0
        status_msg   = await event.respond(f"📢 שולח ל-{len(user_ids)} משתמשים...")
        for uid in user_ids:
            try:
                await bot_client.send_message(uid, f"📢 **הודעה מהמנהל:**\n\n{msg}")
                sent += 1
                await asyncio.sleep(0.05)
            except Exception:
                failed += 1
        await status_msg.edit(f"✅ נשלח ל-{sent} משתמשים. נכשל: {failed}.")

    # ── /quiet ──
    @bot_client.on(events.NewMessage(pattern='/quiet'))
    async def cmd_quiet(event):
        uid      = event.sender_id
        settings = await db_get_user_settings(uid)
        new_val  = 0 if settings['quiet'] else 1
        await db_set_user_setting(uid, 'quiet', new_val)
        status = "מופעל 🔕" if new_val else "כבוי 🔔"
        await event.respond(f"מצב שקט: **{status}**")

    # ── /lang ──
    @bot_client.on(events.NewMessage(pattern=r'/lang ([a-z]{2})'))
    async def cmd_lang(event):
        lang = event.pattern_match.group(1)
        await db_set_user_setting(event.sender_id, 'lang', lang)
        flag = LANGUAGE_MAP.get(lang, '🌐')
        await event.respond(f"שפת ממשק שונתה ל: {flag} `{lang}`")

    # ── /cancel ──
    @bot_client.on(events.NewMessage(pattern='/cancel'))
    async def cmd_cancel(event):
        uid = event.sender_id
        if uid in active_searches:
            active_searches.discard(uid)
            await event.respond("⛔ החיפוש בוטל.")
        else:
            await event.respond("אין חיפוש פעיל כרגע.")

    # ── /subscribe / /unsubscribe ──
    @bot_client.on(events.NewMessage(pattern=r'/subscribe (.+)'))
    async def cmd_subscribe(event):
        query   = event.pattern_match.group(1).strip()
        results = await full_search(user_client, query)
        await db_subscribe(event.sender_id, query, len(results))
        await event.respond(f"🔔 נרשמת לעדכונים עבור: `{query}`\nיש כעת {len(results)} תוצאות.")

    @bot_client.on(events.NewMessage(pattern=r'/unsubscribe (.+)'))
    async def cmd_unsubscribe(event):
        query = event.pattern_match.group(1).strip()
        await db_unsubscribe(event.sender_id, query)
        await event.respond(f"✅ בוטלה ההרשמה עבור: `{query}`")

    # ── /filter ──
    @bot_client.on(events.NewMessage(pattern=r'/filter (\d+) (.+)'))
    async def cmd_filter(event):
        await _do_search(
            event,
            event.pattern_match.group(2).strip(),
            min_members=int(event.pattern_match.group(1))
        )

    # ── /search / /deep ──
    @bot_client.on(events.NewMessage(pattern=r'/search (.+)'))
    async def cmd_search(event):
        await _do_search(event, event.pattern_match.group(1).strip())

    @bot_client.on(events.NewMessage(pattern=r'/deep (.+)'))
    async def cmd_deep(event):
        query = event.pattern_match.group(1).strip()
        await db_clear_cache(query)
        await _do_search(event, query, force_fresh=True)

    # ── /smart ──
    @bot_client.on(events.NewMessage(pattern='/smart'))
    async def cmd_smart(event):
        pending_search[event.sender_id] = {'step': 'query'}
        await event.respond(
            "🧠 **חיפוש חכם**\n\nשלח את מונח החיפוש שלך:",
            buttons=[[Button.inline("❌ ביטול", data="smart_cancel")]]
        )

    @bot_client.on(events.CallbackQuery(data=b"smart_cancel"))
    async def cb_smart_cancel(event):
        pending_search.pop(event.sender_id, None)
        await event.edit("❌ החיפוש האינטרקטיבי בוטל.")

    @bot_client.on(events.CallbackQuery(pattern=rb'smart_members:([^:]+):(\d+)'))
    async def cb_smart_members(event):
        uid   = event.sender_id
        query = event.pattern_match.group(1).decode()
        min_m = int(event.pattern_match.group(2))
        pending_search[uid] = {'step': 'lang', 'query': query, 'min_members': min_m}
        await event.edit(
            f"🌐 **סינון לפי שפה** | שאילתה: `{query}` | מינ': {min_m:,}\n\nבחר שפה:",
            buttons=[
                [Button.inline("🌐 הכל",    data=f"smart_lang:{query}:{min_m}:"),
                 Button.inline("🇮🇱 עברית",  data=f"smart_lang:{query}:{min_m}:he")],
                [Button.inline("🇺🇸 אנגלית", data=f"smart_lang:{query}:{min_m}:en"),
                 Button.inline("🇷🇺 רוסית",  data=f"smart_lang:{query}:{min_m}:ru")],
            ]
        )

    @bot_client.on(events.CallbackQuery(pattern=rb'smart_lang:([^:]+):(\d+):(.*)'))
    async def cb_smart_lang(event):
        uid   = event.sender_id
        query = event.pattern_match.group(1).decode()
        min_m = int(event.pattern_match.group(2))
        lang  = event.pattern_match.group(3).decode()
        pending_search.pop(uid, None)
        await event.edit(f"🔍 מחפש `{query}`...")
        await _do_search(event, query, min_members=min_m, lang_filter=lang)

    # ── /trending ──
    @bot_client.on(events.NewMessage(pattern='/trending'))
    async def cmd_trending(event):
        trending = await db_get_trending(hours=24, limit=10)
        if not trending:
            await event.respond(
                "📭 אין מספיק נתוני צמיחה עדיין.\n"
                "נתונים נצברים אוטומטית עם כל חיפוש."
            )
            return
        lines = [
            f"**{i}. @{t['username']}**\n"
            f"📈 +{t['growth']:,} | 👥 {t['current']:,}\n"
            f"🔗 https://t.me/{t['username']}"
            for i, t in enumerate(trending, 1)
        ]
        await event.respond("🔥 **ערוצים עולים (24h):**\n\n" + "\n\n".join(lines))

    # ── /growth ──
    @bot_client.on(events.NewMessage(pattern=r'/growth @?(\w+)'))
    async def cmd_growth(event):
        username = event.pattern_match.group(1)
        data_7d  = await db_get_growth(username, hours=168)
        data_24h = await db_get_growth(username, hours=24)
        if not data_7d:
            await event.respond(f"📭 אין נתוני צמיחה עבור @{username}.\nחפש אותו קודם.")
            return
        chart      = draw_growth_chart(data_7d)
        change_24h = (data_24h[-1]['members'] - data_24h[0]['members']) if len(data_24h) >= 2 else 0
        change_7d  = (data_7d[-1]['members']  - data_7d[0]['members'])  if len(data_7d)  >= 2 else 0
        sign = lambda n: '+' if n >= 0 else ''
        await event.respond(
            f"📈 **צמיחה: @{username}**\n\n"
            f"24 שעות: **{sign(change_24h)}{change_24h:,}**\n"
            f"7 ימים:  **{sign(change_7d)}{change_7d:,}**\n\n"
            f"```\n{chart}\n```"
        )

    # ── /similar ──
    @bot_client.on(events.NewMessage(pattern=r'/similar @?(\w+)'))
    async def cmd_similar(event):
        username = event.pattern_match.group(1)
        msg      = await event.respond(f"🔗 מחפש ערוצים דומים ל @{username}...")
        try:
            info = await enrich_channel_info(user_client, username)
            if not info:
                await msg.edit(f"❌ לא נמצא @{username}")
                return
            search_term = info['title']
            if info.get('about'):
                search_term += ' ' + ' '.join(info['about'].split()[:3])
            results = await full_search(user_client, search_term)
            results = [r for r in results if r['username'] != username][:15]
            await msg.delete()
            if not results:
                await event.respond("❌ לא נמצאו ערוצים דומים")
                return
            await _send_page(event, f"similar:{username}", results, 0)
        except Exception as e:
            await msg.edit(f"❌ שגיאה: {e}")

    # ── /preview ──
    @bot_client.on(events.NewMessage(pattern=r'/preview @?(\w+)'))
    async def cmd_preview(event):
        username = event.pattern_match.group(1)
        msg      = await event.respond(f"👁 טוען תצוגה מקדימה של @{username}...")
        try:
            entity = await user_client(ResolveUsernameRequest(username))
            if not hasattr(entity, 'chats') or not entity.chats:
                await msg.edit(f"❌ @{username} לא נמצא")
                return
            chat  = entity.chats[0]
            msgs  = await get_recent_messages(user_client, chat, limit=5)
            text  = f"👁 **{chat.title}** | t.me/{username}\n\n"
            text += ("**5 הודעות אחרונות:**\n" + "\n".join(msgs)) if msgs else "_לא נמצאו הודעות ציבוריות_"
            await msg.edit(text)
        except Exception as e:
            await msg.edit(f"❌ שגיאה: {e}")

    # ── Inline Mode ──
    @bot_client.on(events.InlineQuery)
    async def cb_inline(event):
        query = event.text.strip()
        if len(query) < 2:
            await event.answer([])
            return
        cached = await db_get_cache(query)
        if not cached:
            await event.answer(
                results=[],
                switch_pm='חיפוש מלא — לחץ כאן',
                switch_pm_param='inline_start'
            )
            return
        articles = [
            event.builder.article(
                title=f"{r['title']} ❤️{r.get('health',0)} ({r['members']:,})",
                description=r.get('about', '')[:50],
                text=f"**{r['title']}**\n👥 {r['members']:,} | ❤️ {r.get('health',0)}/100\n🔗 {r['link']}"
            )
            for r in cached[:10]
        ]
        await event.answer(articles)

    # ── טקסט חופשי ──
    @bot_client.on(events.NewMessage(incoming=True))
    async def cmd_free_text(event):
        uid = event.sender_id
        if uid in pending_search:
            state = pending_search[uid]
            if state.get('step') == 'query':
                text = event.text.strip()
                if not text or text.startswith('/'):
                    return
                pending_search[uid] = {'step': 'members', 'query': text}
                await event.respond(
                    f"📊 **סף מנויים מינימלי** (שאילתה: `{text}`)\n\nבחר:",
                    buttons=[
                        [Button.inline("0 (הכל)",  data=f"smart_members:{text}:0"),
                         Button.inline("100+",      data=f"smart_members:{text}:100")],
                        [Button.inline("1,000+",   data=f"smart_members:{text}:1000"),
                         Button.inline("10,000+",  data=f"smart_members:{text}:10000")],
                    ]
                )
                return
        if event.is_private and not event.text.startswith('/'):
            await _do_search(event, event.text.strip())

    # ── Callbacks ──
    @bot_client.on(events.CallbackQuery(pattern=rb'page:(.+):(\d+)'))
    async def cb_page(event):
        query   = event.pattern_match.group(1).decode()
        page    = int(event.pattern_match.group(2))
        results = await db_get_cache(query) or []
        await _send_page(event, query, results, page, edit=True)

    @bot_client.on(events.CallbackQuery(pattern=rb'refresh:(.+)'))
    async def cb_refresh(event):
        query = event.pattern_match.group(1).decode()
        await db_clear_cache(query)
        await event.answer("🔄 מרענן...")
        results = await full_search(user_client, query, force_fresh=True)
        await _send_page(event, query, results, 0, edit=True)

    @bot_client.on(events.CallbackQuery(pattern=rb'sub:(.+)'))
    async def cb_sub(event):
        query   = event.pattern_match.group(1).decode()
        results = await db_get_cache(query) or []
        await db_subscribe(event.sender_id, query, len(results))
        await event.answer(f"🔔 נרשמת ל: {query}")

    @bot_client.on(events.CallbackQuery(pattern=rb'share:(.+)'))
    async def cb_share(event):
        query   = event.pattern_match.group(1).decode()
        results = await db_get_cache(query) or []
        if not results:
            await event.answer("❌ אין תוצאות במטמון")
            return
        top3 = results[:3]
        text = f"🔍 **תוצאות חיפוש: {query}**\n\n"
        for i, r in enumerate(top3, 1):
            text += f"{i}. **{r['title']}** | 👥 {r['members']:,}\n🔗 {r['link']}\n\n"
        text += f"_ועוד {len(results)-3} תוצאות — חפש בבוט_" if len(results) > 3 else ""
        await bot_client.send_message(event.chat_id, text, link_preview=False)
        await event.answer("📤 נשלח!")

    @bot_client.on(events.CallbackQuery(pattern=rb'csv:(.+)'))
    async def cb_csv(event):
        query   = event.pattern_match.group(1).decode()
        results = await db_get_cache(query) or []
        if not results:
            await event.answer("❌ אין תוצאות במטמון")
            return
        await bot_client.send_file(
            event.chat_id,
            file=io.BytesIO(results_to_csv(results)),
            attributes=[types.DocumentAttributeFilename(file_name=f"{query}_results.csv")],
            caption=f"📥 **{len(results)} תוצאות:** `{query}`"
        )
        await event.answer("📥 CSV נשלח!")

    @bot_client.on(events.CallbackQuery(pattern=rb'xlsx:(.+)'))
    async def cb_xlsx(event):
        query   = event.pattern_match.group(1).decode()
        results = await db_get_cache(query) or []
        if not results:
            await event.answer("❌ אין תוצאות במטמון")
            return
        await event.answer("📊 מייצר Excel...")
        ext  = 'xlsx' if HAS_OPENPYXL else 'csv'
        await bot_client.send_file(
            event.chat_id,
            file=io.BytesIO(results_to_xlsx(results)),
            attributes=[types.DocumentAttributeFilename(file_name=f"{query}_results.{ext}")],
            caption=f"📊 **{len(results)} תוצאות:** `{query}`"
        )

    @bot_client.on(events.CallbackQuery(pattern=rb'pdf:(.+)'))
    async def cb_pdf(event):
        query   = event.pattern_match.group(1).decode()
        results = await db_get_cache(query) or []
        if not results:
            await event.answer("❌ אין תוצאות במטמון")
            return
        await event.answer("📄 מייצר PDF...")
        data = results_to_pdf(results, query)
        ext  = 'pdf' if HAS_FPDF else 'csv'
        await bot_client.send_file(
            event.chat_id,
            file=io.BytesIO(data),
            attributes=[types.DocumentAttributeFilename(file_name=f"{query}_results.{ext}")],
            caption=f"📄 **{len(results)} תוצאות:** `{query}`"
        )

    # ── שליחת עמוד ──
    async def _send_page(event, query, results, page, edit=False):
        start = page * 10
        chunk = results[start:start + 10]
        if not chunk:
            await event.answer("❌ אין עוד תוצאות")
            return
        total_pages = (len(results) - 1) // 10 + 1
        text = (
            f"📄 עמוד {page + 1}/{total_pages} | {len(results)} תוצאות עבור `{query}`\n\n"
            + "\n\n".join(format_result(r, start + i + 1) for i, r in enumerate(chunk))
        )
        buttons = make_buttons(query, page, len(results))
        try:
            if edit:
                await event.edit(text, buttons=buttons, link_preview=False)
            else:
                await bot_client.send_message(event.chat_id, text, buttons=buttons, link_preview=False)
        except Exception:
            await bot_client.send_message(event.chat_id, text, buttons=buttons, link_preview=False)

    # ── חיפוש מרכזי ──
    async def _do_search(
        event,
        query:        str,
        min_members:  int  = 0,
        force_fresh:  bool = False,
        lang_filter:  str  = '',
        cat_filter:   str  = '',
        custom_title: str  = '',
    ):
        uid      = event.sender_id
        settings = await db_get_user_settings(uid)

        if ALLOWED_USERS and uid not in ALLOWED_USERS:
            await event.respond("⛔ אין לך הרשאה.")
            return
        if not await _check_terms(event):
            return
        if not await db_check_rate(uid):
            await event.respond(f"⏳ הגעת למגבלת {RATE_LIMIT_PER_HOUR} חיפושים לשעה.")
            return

        fixed = fix_typo(query)
        if fixed != query:
            await event.respond(f"✏️ תיקון: `{query}` → `{fixed}`")
            query = fixed

        title      = custom_title or query
        status_msg = await event.respond(f"🌐 מחפש: **{title}**\n⏳ אנא המתן...")
        active_searches.add(uid)

        async def progress_cb(text: str):
            if uid not in active_searches:
                raise asyncio.CancelledError()
            try:
                await status_msg.edit(text)
            except Exception:
                pass

        try:
            results = await full_search(
                user_client, query,
                min_members=min_members,
                force_fresh=force_fresh,
                lang_filter=lang_filter,
                cat_filter=cat_filter,
                progress_cb=progress_cb,
            )
        except asyncio.CancelledError:
            await status_msg.edit("⛔ החיפוש בוטל.")
            return
        except Exception as e:
            log.error(f"Search error: {e}\n{traceback.format_exc()}")
            await status_msg.edit("❌ שגיאה בחיפוש. נסה שוב.")
            for admin in ADMIN_IDS:
                try:
                    await bot_client.send_message(
                        admin,
                        f"⚠️ **שגיאת חיפוש:**\n`{e}`\nמשתמש: {uid}\nשאילתה: {query}"
                    )
                except Exception:
                    pass
            return
        finally:
            active_searches.discard(uid)

        await status_msg.delete()

        if not results:
            await event.respond(f"❌ לא נמצאו קבוצות עבור **{title}**")
            return

        if settings['quiet']:
            ext  = 'xlsx' if HAS_OPENPYXL else 'csv'
            await bot_client.send_file(
                event.chat_id,
                file=io.BytesIO(results_to_xlsx(results)),
                attributes=[types.DocumentAttributeFilename(file_name=f"{query}_results.{ext}")],
                caption=f"📊 **{len(results)} תוצאות:** `{title}` (מצב שקט)"
            )
            await db_log_search(uid, query, len(results))
            return

        await _send_page(event, query, results, 0)
        await db_log_search(uid, query, len(results))
        log.info(f"Done: '{query}' → {len(results)} results for {uid}")

# ─────────────────────────────────────────────
#  משימות רקע
# ─────────────────────────────────────────────

async def subscription_checker(bot_client, user_client):
    while True:
        await asyncio.sleep(3600)
        log.info("Checking subscriptions...")
        for sub in await db_get_subscriptions():
            try:
                results   = await full_search(user_client, sub['query'])
                new_count = len(results)
                if new_count > sub['last_count']:
                    diff = new_count - sub['last_count']
                    await bot_client.send_message(
                        sub['user_id'],
                        f"🔔 **עדכון מנוי:** `{sub['query']}`\n"
                        f"נוספו **{diff}** ערוצים חדשים! סה\"כ: {new_count}\n"
                        f"שלח `/search {sub['query']}` לצפייה"
                    )
                    await db_subscribe(sub['user_id'], sub['query'], new_count)
                await asyncio.sleep(2)
            except Exception as e:
                log.warning(f"Subscription error: {e}")

async def alert_checker(bot_client, user_client):
    while True:
        await asyncio.sleep(1800)  # כל 30 דקות
        log.info("Checking alerts...")
        for alert in await db_get_all_alerts():
            try:
                results  = await full_search(user_client, alert['keyword'])
                new_uids = {r['username'] for r in results}
                old_uids = set(alert['last_found'])
                added    = new_uids - old_uids

                if added:
                    new_channels = [r for r in results if r['username'] in added][:5]
                    lines = [
                        f"• **{r['title']}** | 👥 {r['members']:,} | ❤️{r.get('health',0)}\n  {r['link']}"
                        for r in new_channels
                    ]
                    await bot_client.send_message(
                        alert['user_id'],
                        f"🔔 **Alert: `{alert['keyword']}`**\n"
                        f"נמצאו **{len(added)}** ערוצים חדשים!\n\n" + "\n\n".join(lines)
                    )
                    await db_update_alert_found(
                        alert['user_id'], alert['keyword'], list(new_uids)
                    )
                await asyncio.sleep(1)
            except Exception as e:
                log.warning(f"Alert error: {e}")

async def daily_backup(bot_client):
    while True:
        await asyncio.sleep(86400)
        if not ADMIN_IDS or not os.path.exists(DB_PATH):
            continue
        log.info("Sending daily backup...")
        for admin in ADMIN_IDS:
            try:
                await bot_client.send_file(
                    admin, file=DB_PATH,
                    caption=f"💾 **גיבוי יומי** — {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
                )
            except Exception as e:
                log.warning(f"Backup error: {e}")

async def watchdog(bot_client, user_client):
    while True:
        await asyncio.sleep(60)
        try:
            await user_client.get_me()
        except Exception as e:
            log.error(f"Watchdog: user_client disconnected: {e}")
            for admin in ADMIN_IDS:
                try:
                    await bot_client.send_message(
                        admin, f"⚠️ **Watchdog:** user_client התנתק!\n`{e}`"
                    )
                except Exception:
                    pass

# ─────────────────────────────────────────────
#  הרצה
# ─────────────────────────────────────────────

async def run_all():
    global ENRICH_SEM
    ENRICH_SEM = asyncio.Semaphore(5)  # חייב להיווצר בתוך event loop (Termux / Python 3.10+)

    await init_db()
    await _get_redis()  # נסה לחבר Redis

    bot_client  = TelegramClient('bot_session',  API_ID, API_HASH)
    user_client = TelegramClient('user_session', API_ID, API_HASH)

    await bot_client.start(bot_token=BOT_TOKEN)
    await user_client.start()

    register_handlers(bot_client, user_client)

    for admin in ADMIN_IDS:
        try:
            await bot_client.send_message(
                admin,
                f"🚀 **הבוט עלה! v3.0**\n"
                f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC\n"
                f"📊 Excel: {'✅' if HAS_OPENPYXL else '❌'} | "
                f"📄 PDF: {'✅' if HAS_FPDF else '❌'} | "
                f"🔴 Redis: {'✅' if HAS_REDIS else '❌'}\n"
                f"🔀 Proxies: {len(PROXY_LIST)}"
            )
        except Exception:
            pass

    asyncio.create_task(subscription_checker(bot_client, user_client))
    asyncio.create_task(alert_checker(bot_client, user_client))
    asyncio.create_task(daily_backup(bot_client))
    asyncio.create_task(watchdog(bot_client, user_client))

    log.info("🚀 Ultra Bot v3.0 running!")
    await bot_client.run_until_disconnected()


if __name__ == '__main__':
    try:
        import sys
        # Termux / Android: אין uvloop, אבל asyncio רגיל עובד
        if sys.platform == 'linux' and 'TERMUX_VERSION' in os.environ:
            log.info("Termux detected — using standard asyncio event loop")
        try:
            asyncio.run(run_all())
        except AttributeError:
            # Python < 3.7 fallback (נדיר, אבל בטוח)
            loop = asyncio.get_event_loop()
            loop.run_until_complete(run_all())
    except KeyboardInterrupt:
        log.info("Bot stopped by user.")
    except Exception as e:
        log.critical(f"Fatal error: {e}\n{traceback.format_exc()}")
